import sqlite3
from openpyxl import load_workbook



excel_file=load_workbook("C:/Users/min17/Desktop/대주가능목록/20221226 대주가능목록.xlsx", data_only=True)
datal=excel_file['Sheet1']

conn=sqlite3.connect("Daum_Crawling_list.db",isolation_level=None)
cur=conn.cursor()

row=2

while datal.cell(row,1).value is not None:
    cur.execute("insert into 대주가능목록(code,name,class,price,quantity) values (?,?,?,?,?)",
                (datal.cell(row,1).value)
                (datal.cell(row,2).value)
                (datal.cell(row,3).value)
                (datal.cell(row,4).value
                (datal.cell(row,5).value))
                )
    row=row+1


